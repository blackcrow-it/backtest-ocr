import pandas as pd
from difflib import SequenceMatcher
from api import check_ocr_image_url_fpt
from helper import percent_sequence_matcher
import base64
import requests
import datetime
import time

from openpyxl import Workbook
from openpyxl.styles import colors, Color, Font, PatternFill
from datetime import datetime

def read_file_excel(file, sheet):
  df = pd.read_excel(file, sheet, engine='openpyxl', dtype = str)
  return df.iterrows()

def write_info_to_excel(data):
  wb = Workbook()
  try:
    ws = wb[str(datetime.now().strftime("%d-%m-%Y %H.%M.%S"))]
  except KeyError:
    print(KeyError)
    ws = wb.create_sheet(str(datetime.now().strftime("%d-%m-%Y %H.%M.%S")))
  ws['A1'] = 'PawnID'
  ws['B1'] = 'Tên khách hàng'
  ws['C1'] = 'Tên chủ xe'
  ws['D1'] = 'Biển số đăng ký'
  ws['E1'] = 'Số khung'
  ws['F1'] = 'Số máy'
  ws['G1'] = 'Số thẻ'
  ws['H1'] = 'Nhãn hiệu'
  ws['I1'] = 'Số loại'
  ws['J1'] = 'Màu sơn'
  ws['K1'] = 'Loại xe'
  ws['L1'] = 'Dung tích'
  ws['M1'] = 'Số chỗ ngồi'
  ws['N1'] = 'Tải trọng'
  ws['O1'] = 'Tải trọng: Hàng hóa'
  ws['P1'] = 'Năm sản xuất'
  ws['Q1'] = 'Đăng ký lần đầu ngày'
  ws['R1'] = 'Đăng ký xe có giá trị đến ngày'
  ws['S1'] = 'Địa chỉ'
  ws['T1'] = 'Thời gian OCR'
  ws['U1'] = 'Ảnh thường'
  ws['V1'] = 'Tên (score)'
  ws['W1'] = 'Biển số (score)'
  ws['X1'] = 'Số khung (score)'
  ws['Y1'] = 'Số máy (score)'
  ws['Z1'] = 'Số thẻ (score)'
  ws['AA1'] = 'Ảnh 1'
  ws['AB1'] = 'Ảnh 2'
  wb.save('result/fpt-dkx.xlsx')

  for index, item in data:
    print(item['CodeNo'] + '-' + item['PawnID'])
    ws['A' + str(index + 2)] = item['PawnID']
    ws['B' + str(index + 2)] = item['Tên khách hàng']
    is_normal_image = True
    try:
      result, time_response = check_ocr_image_url_fpt(item['Image2'], item['Image1'], 'dkx')
      ws['AA' + str(index + 2)] = item['Image1']
      ws['AB' + str(index + 2)] = item['Image2']
    except:
      try:
        is_normal_image = False
        result, time_response = check_ocr_image_url_fpt(item['Image3'], item['Image4'], 'dkx')
        ws['AA' + str(index + 2)] = item['Image3']
        ws['AB' + str(index + 2)] = item['Image4']
      except:
        continue

    if result['status_code'] == 200:
      data_dkx = result['data']
      try:
        ws['C' + str(index + 2)] = data_dkx['hoVaTen']
        ws['D' + str(index + 2)] = data_dkx['bienSo']
        ws['E' + str(index + 2)] = data_dkx['soKhung']
        ws['F' + str(index + 2)] = data_dkx['soMay']
        ws['G' + str(index + 2)] = data_dkx['soChungNhan'] if 'soChungNhan' in data_dkx else ''
        ws['H' + str(index + 2)] = data_dkx['nhanHieu']
        ws['I' + str(index + 2)] = data_dkx['soLoai']
        ws['J' + str(index + 2)] = data_dkx['mauSon']
        ws['K' + str(index + 2)] = data_dkx['loaiXe']
        ws['L' + str(index + 2)] = data_dkx['dungTich']
        ws['M' + str(index + 2)] = data_dkx['soNguoi']
        ws['N' + str(index + 2)] = data_dkx['taiTrong']
        ws['O' + str(index + 2)] = ''
        ws['P' + str(index + 2)] = ''
        ws['Q' + str(index + 2)] = data_dkx['ngayDangKyLanDau']
        ws['R' + str(index + 2)] = ''
        ws['S' + str(index + 2)] = data_dkx['diaChi']
        ws['V' + str(index + 2)] = percent_sequence_matcher(item['Tên khách hàng'], data_dkx['hoVaTen'])
        ws['W' + str(index + 2)] = percent_sequence_matcher(item['BKS'], data_dkx['bienSo'])
        ws['X' + str(index + 2)] = percent_sequence_matcher(item['Số khung'], data_dkx['soKhung'])
        ws['Y' + str(index + 2)] = percent_sequence_matcher(item['Số máy'], data_dkx['soMay'])
        ws['Z' + str(index + 2)] = percent_sequence_matcher(item['Số thẻ'], data_dkx['soChungNhan']) if 'soChungNhan' in data_dkx else 0
      except:
          pass
    ws['T' + str(index + 2)] = time_response
    if is_normal_image:
      ws['U' + str(index + 2)] = 'x'
    wb.save('result/fpt-dkx.xlsx')
    time.sleep(5)

def get_as_base64(url):
  return base64.b64encode(requests.get(url).content)

def main():
  data = read_file_excel('./map-dkx.xlsx', '10-10-2021 11.23.56')
  write_info_to_excel(data)

if __name__ == '__main__':
    main()
    