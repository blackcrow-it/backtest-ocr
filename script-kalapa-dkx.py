import pandas as pd
from difflib import SequenceMatcher
from api import check_ocr_image_url_kalapa
from helper import percent_sequence_matcher
import base64
import requests
import datetime
import time

from openpyxl import Workbook
from openpyxl.styles import colors, Color, Font, PatternFill
from datetime import datetime

def read_file_excel(file, sheet):
  df = pd.read_excel(file, sheet, engine='openpyxl', dtype = str)
  return df.iterrows()

def write_info_to_excel(data):
  wb = Workbook()
  try:
    ws = wb[str(datetime.now().strftime("%d-%m-%Y %H.%M.%S"))]
  except KeyError:
    print(KeyError)
    ws = wb.create_sheet(str(datetime.now().strftime("%d-%m-%Y %H.%M.%S")))
  ws['A1'] = 'PawnID'
  ws['B1'] = 'Tên khách hàng'
  ws['C1'] = 'Tên chủ xe'
  ws['D1'] = 'Biển số đăng ký'
  ws['E1'] = 'Số khung'
  ws['F1'] = 'Số máy'
  ws['G1'] = 'Số thẻ'
  ws['H1'] = 'Nhãn hiệu'
  ws['I1'] = 'Số loại'
  ws['J1'] = 'Màu sơn'
  ws['K1'] = 'Loại xe'
  ws['L1'] = 'Dung tích'
  ws['M1'] = 'Số chỗ ngồi'
  ws['N1'] = 'Tải trọng'
  ws['O1'] = 'Tải trọng: Hàng hóa'
  ws['P1'] = 'Năm sản xuất'
  ws['Q1'] = 'Đăng ký lần đầu ngày'
  ws['R1'] = 'Đăng ký xe có giá trị đến ngày'
  ws['S1'] = 'Địa chỉ'
  ws['T1'] = 'Thời gian OCR'
  ws['U1'] = 'Ảnh thường'
  ws['V1'] = 'Tên (score)'
  ws['W1'] = 'Biển số (score)'
  ws['X1'] = 'Số khung (score)'
  ws['Y1'] = 'Số máy (score)'
  ws['Z1'] = 'Số thẻ (score)'
  ws['AA1'] = 'Ảnh'
  wb.save('result/fpt-dkx.xlsx')

  for index, item in data:
    print(item['CodeNo'] + '-' + item['PawnID'])
    ws['A' + str(index + 2)] = item['PawnID']
    ws['B' + str(index + 2)] = item['Tên khách hàng']
    if (item['Image1'] == ''):
      continue
    is_normal_image = True
    try:
      result, time_response = check_ocr_image_url_kalapa(item['Image1'])
      ws['AA' + str(index + 2)] = item['Image1']
    except:
      continue

    if result['status_code'] == 200:
      data_dkx = result['data']
      try:
        ws['C' + str(index + 2)] = data_dkx['name']
        ws['D' + str(index + 2)] = data_dkx['plate']
        ws['E' + str(index + 2)] = data_dkx['chassis']
        ws['F' + str(index + 2)] = data_dkx['engine']
        ws['G' + str(index + 2)] = ''
        ws['H' + str(index + 2)] = data_dkx['brand']
        ws['I' + str(index + 2)] = data_dkx['model']
        ws['J' + str(index + 2)] = data_dkx['color']
        ws['K' + str(index + 2)] = ''
        ws['L' + str(index + 2)] = data_dkx['capacity']
        ws['M' + str(index + 2)] = ''
        ws['N' + str(index + 2)] = ''
        ws['O' + str(index + 2)] = ''
        ws['P' + str(index + 2)] = ''
        ws['Q' + str(index + 2)] = data_dkx['firstRegistration']
        ws['R' + str(index + 2)] = ''
        ws['S' + str(index + 2)] = data_dkx['address']
        ws['V' + str(index + 2)] = percent_sequence_matcher(item['Tên khách hàng'], data_dkx['name'])
        ws['W' + str(index + 2)] = percent_sequence_matcher(item['BKS'], data_dkx['plate'])
        ws['X' + str(index + 2)] = percent_sequence_matcher(item['Số khung'], data_dkx['chassis'])
        ws['Y' + str(index + 2)] = percent_sequence_matcher(item['Số máy'], data_dkx['engine'])
      except:
          pass
    ws['T' + str(index + 2)] = time_response
    if is_normal_image:
      ws['U' + str(index + 2)] = 'x'
    wb.save('result/kalapa-dkx.xlsx')

def get_as_base64(url):
  return base64.b64encode(requests.get(url).content)

def main():
  data = read_file_excel('./map-dkx.xlsx', '10-10-2021 11.23.56')
  write_info_to_excel(data)

if __name__ == '__main__':
    main()
    