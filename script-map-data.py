import pandas as pd
import xlsxwriter
from difflib import SequenceMatcher
from api import check_ocr_dkx_image_url_gmo
import base64
import requests

from openpyxl import Workbook
from datetime import datetime

df = pd.read_excel("./ca-vet-xe.xlsx", "Sheet1", engine='openpyxl', dtype = str)
data = {}
for index, row in df.iterrows():
  item = row.to_dict()
  if item['CodeNo'] in data:
    data[item['CodeNo']]['Image'].append(item['Image'])
  else:
    item['Image'] = [item['Image']]
    data[item['CodeNo']] = item

wb = Workbook()
try:
  ws = wb[str(datetime.now().strftime("%d-%m-%Y %H.%M.%S"))]
except KeyError:
  print(KeyError)
  ws = wb.create_sheet(str(datetime.now().strftime("%d-%m-%Y %H.%M.%S")))
ws['A1'] = 'CodeNo'
ws['B1'] = 'PawnID'
ws['C1'] = 'BKS'
ws['D1'] = 'Số khung'
ws['E1'] = 'Số máy'
ws['F1'] = 'Năm sản xuất'
ws['I1'] = 'Tên khách hàng'
ws['J1'] = 'Image1'
ws['K1'] = 'Image2'
ws['L1'] = 'Image3'
ws['M1'] = 'Image4'
index = 2
for key, value in data.items():
  print(value['CodeNo'])
  ws['A' + str(index)] = value['CodeNo']
  ws['B' + str(index)] = value['PawnID']
  ws['C' + str(index)] = value['BKS']
  ws['D' + str(index)] = value['Số khung']
  ws['E' + str(index)] = value['Số máy']
  ws['F' + str(index)] = value['Năm sản xuất']
  ws['I' + str(index)] = value['Tên khách hàng']
  ws['J' + str(index)] = value['Image'][0]
  try:
    ws['K' + str(index)] = value['Image'][1]
  except:
    pass
  try:
    ws['L' + str(index)] = value['Image'][2]
  except:
    pass
  try:
    ws['M' + str(index)] = value['Image'][3]
  except:
    pass
  wb.save('map-dks.xlsx')
  index += 1
  
# print(data)